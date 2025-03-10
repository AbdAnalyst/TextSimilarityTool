"""
Text Similarity Analysis Script

Author: Abdullah Choudhry
Repository: https://github.com/AbdAnalyst/TextSimilarityTool

Purpose:
This script analyzes text similarity by:
- Removing stopwords and punctuation.
- Tokenizing and comparing texts using a bag-of-words approach.
- Calculating similarity percentages.
- Highlighting common words in HTML.
- Saving only the first 20 results in both HTML and Excel.

Usage:
1. Install dependencies:
   pip install pandas openpyxl nltk ipython
2. Download NLTK stopwords:
   import nltk
   nltk.download('stopwords')
3. Place the dataset (Excel file) in the same folder.
4. Run the script:
   python similarity_analysis.py
"""

import re
import pandas as pd
from IPython.display import display, HTML, FileLink
import openpyxl
from openpyxl.styles import Font
from openpyxl.cell.rich_text import CellRichText, TextBlock

import nltk
from nltk.corpus import stopwords

# 1) Load Data
stop_words = set(stopwords.words('english'))
df = pd.read_excel("VariedRiskAssessmentDataset.xlsx", engine="openpyxl")
col = "describeTheSpecificHazards"  # The column to analyze
text_data = df[col].dropna().tolist()[:20]  # Only take first 20 rows

# 2) Clean & Tokenize
def clean_tokenize(txt):
    """
    Lowercases text, removes punctuation and stopwords,
    returns a set of valid alphanumeric tokens.
    """
    txt = txt.lower()
    tokens = re.findall(r"[a-z0-9]+", txt)
    return {t for t in tokens if t not in stop_words}

# 3) Calculate Overlap
def bow_sim(txt1, txt2):
    """
    Returns overlap percentage and the set of common tokens.
    """
    w1 = clean_tokenize(txt1)
    w2 = clean_tokenize(txt2)
    common = w1.intersection(w2)
    avg_size = (len(w1) + len(w2)) / 2.0
    if avg_size == 0:
        return 0.0, set()
    return (len(common) / avg_size) * 100, common

# 4) Highlight Common Words in HTML
def highlight_common(txt, common_set, color="#FFD54F"):
    """
    Highlights tokens in 'txt' if they appear in 'common_set'.
    Returns an HTML string with highlighted text.
    """
    orig_tokens = re.findall(r"\S+", txt)
    highlighted = []
    for token in orig_tokens:
        cleaned_parts = re.findall(r"[a-z0-9]+", token.lower())
        if any(cp in common_set for cp in cleaned_parts):
            highlighted.append(f"<span style='background-color:{color}'>{token}</span>")
        else:
            highlighted.append(token)
    return " ".join(highlighted)

# 5) Compare Each Pair & Collect Results
results = []
for i in range(len(text_data)):
    for j in range(i+1, len(text_data)):
        t1 = text_data[i]
        t2 = text_data[j]
        similarity, common_w = bow_sim(t1, t2)
        
        highlighted_1 = highlight_common(t1, common_w, color="#FFD54F")
        highlighted_2 = highlight_common(t2, common_w, color="#81C784")
        
        results.append((t1, t2, similarity, sorted(common_w), highlighted_1, highlighted_2))

# Keep only first 20 comparison results
results = results[:20]

# Build HTML Table
html_rows = []
for (txt1, txt2, sim, common, h1, h2) in results:
    common_str = ", ".join(common) if common else "None"
    row_html = f"""
    <tr>
      <td style="vertical-align:top;">{txt1}</td>
      <td style="vertical-align:top;">{txt2}</td>
      <td style="vertical-align:top; text-align:center;">{sim:.2f}%</td>
      <td style="vertical-align:top;">{common_str}</td>
      <td style="vertical-align:top;">{h1}</td>
      <td style="vertical-align:top;">{h2}</td>
    </tr>
    """
    html_rows.append(row_html)

html_table = f"""
<h2>Overlap (No Stopwords or Punctuation)</h2>
<table border="1" style="border-collapse:collapse;width:100%;">
  <thead>
    <tr>
      <th>Text 1</th>
      <th>Text 2</th>
      <th>Overlap %</th>
      <th>Common Words</th>
      <th>Highlighted Text 1</th>
      <th>Highlighted Text 2</th>
    </tr>
  </thead>
  <tbody>
    {''.join(html_rows)}
  </tbody>
</table>
"""

# Display HTML in Jupyter (if run interactively)
try:
    display(HTML(html_table))
except:
    pass

# Write HTML to file
out_html = "similarity_results.html"
with open(out_html, "w", encoding="utf-8") as f:
    f.write(html_table)
try:
    display(FileLink(out_html))
except:
    pass

print(f"HTML results saved to '{out_html}'.")

# 6) Create an Excel file with partial font-color highlights
def build_rich_text(txt, common_set, highlight_color="FFFF0000"):
    from openpyxl.cell.rich_text import CellRichText, TextBlock
    from openpyxl.cell.text import InlineFont
    
    tokens = re.findall(r"\S+", txt)
    blocks = []
    for i, token in enumerate(tokens):
        space = " " if i < len(tokens) - 1 else ""
        cleaned_parts = re.findall(r"[a-z0-9]+", token.lower())
        if any(cp in common_set for cp in cleaned_parts):
            # highlight token in red
            blocks.append(TextBlock(InlineFont(color=highlight_color), token + space))
        else:
            blocks.append(TextBlock(InlineFont(color="FF000000"), token + space))
    return CellRichText(blocks)

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Similarity Results"
ws.append([
    "Pair #", "Overlap %", "Common Words",
    "Text 1", "Text 2", "Highlighted 1", "Highlighted 2"
])

pair_num = 1
for (txt1, txt2, sim, common, _, _) in results:
    row_idx = ws.max_row + 1
    sim_str = f"{sim:.2f}%"
    common_str = ", ".join(common) if common else "None"
    
    # Add plain text & overlap
    ws.cell(row=row_idx, column=1).value = f"Pair {pair_num}"
    ws.cell(row=row_idx, column=2).value = sim_str
    ws.cell(row=row_idx, column=3).value = common_str
    ws.cell(row=row_idx, column=4).value = txt1
    ws.cell(row=row_idx, column=5).value = txt2
    
    # Build partial-color text
    cell_rich1 = build_rich_text(txt1, set(common))
    cell_rich2 = build_rich_text(txt2, set(common))
    ws.cell(row=row_idx, column=6).value = cell_rich1
    ws.cell(row=row_idx, column=7).value = cell_rich2
    
    pair_num += 1

out_excel = "similarity_results.xlsx"
wb.save(out_excel)
print(f"Excel results saved to '{out_excel}'.")
try:
    display(FileLink(out_excel))
except:
    pass
